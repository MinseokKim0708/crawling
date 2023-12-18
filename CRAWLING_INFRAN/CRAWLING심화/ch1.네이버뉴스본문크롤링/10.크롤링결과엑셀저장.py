import requests
from bs4 import BeautifulSoup
import time
import pyautogui
from openpyxl import Workbook
from openpyxl.styles import Alignment

keyword = pyautogui.prompt("검색어를 입력하세요.") #사용자 검색
lastpage = int(pyautogui.prompt("몇 페이지까지 크롤링 할까요?"))

wb = Workbook() #엑셀 생성
ws = wb.create_sheet(keyword) #엑셀 시트 생성

ws.column_dimensions['A'].width = 60 #열너비변경
ws.column_dimensions['B'].width = 60 #열너비변경
ws.column_dimensions['C'].width = 120 #열너비변경

row = 1 #행 번호
page_num = 1 #페이지 번호
for i in range(1,lastpage*10, 10):
    print(f"{page_num}페이지 크롤링 중 입니다.==================")
    response = requests.get(f'https://search.naver.com/search.naver?where=news&query={keyword}&start={i}')
    html = response.text
    soup = BeautifulSoup(html, 'html.parser')

    articles = soup.select('div.info_group') #뉴스 기사 10개 추출

    for article in articles:
        links = article.select('a.info')
        if len(links) >= 2: #링크가 2개 이상이면
            url = links[1].attrs['href'] #두번째 링크의 href를 추출
            response = requests.get(url, headers={'User-agent':'Mozila/5.0'})
            html = response.text
            soup = BeautifulSoup(html, "html.parser")
        
            if "entertain" in response.url: #만약 연예 뉴스 라면
                title = soup.select_one(".end_tit") 
                content = soup.select_one("#articeBody") #연예 뉴스 본문 내용
            elif "sports" in response.url: #만약 스포츠 뉴스 라면
                title = soup.select_one("h4.title") 
                content = soup.select_one("#newsEndContents") #스포츠 뉴스 본문 내용
                divs = content.select("div") #본문 내용안에 불필요한 내용 삭제
                for div in divs:
                    div.decompose()
                paragraphs = content.select("p")
                for p in paragraphs:
                    p.decompose()
            else:
                title = soup.select_one("#title_area")
                content = soup.select_one('#newsct_article') #기사 본문 내용
            print("==========링크==========\n", url)
            print("==========제목==========\n", title.text.strip())
            print("==========본문==========\n", content.text.strip() )

            ws[f'A{row}'] = url
            ws[f'B{row}'] = title.text.strip()
            ws[f'C{row}'] = content.text.strip()
            ws[f'C{row}'].alignment = Alignment(wrap_text=True) #자동 줄 바꿈
            row += 1
            time.sleep(0.3)
    page_num += 1

wb.save(f'{keyword}_result.xlsx')